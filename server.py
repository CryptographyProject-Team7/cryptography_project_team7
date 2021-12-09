import socket
import threading
import habsd
import timeit
import HashingAndSalting
import hashlib
from datetime import datetime

import pylab
import openpyxl
HEADER = 2048
PORT = 5050
SERVER = socket.gethostbyname(socket.gethostname())
ADDR = (SERVER, PORT)
FORMAT = 'utf-8'
DISCONNECT_MESSAGE = "!DISCONNECT"

server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
server.bind(ADDR)




def handle_client(conn, addr):
    print(f"[NEW CONNECTION] {addr} connected succesfully.")

    connected = True
    while connected:
        msg_length = conn.recv(HEADER).decode(FORMAT)
        if msg_length:
            msg_length = int(msg_length)
            msg = conn.recv(msg_length).decode(FORMAT)
            if msg == DISCONNECT_MESSAGE:
                connected = False




            print(f"[{addr}] {msg}")
            tic=timeit.default_timer()
            encrypted_msg = habsd.encrypt(msg,threading.active_count())
            toc=timeit.default_timer()
            print("\nYour encrypted message is: ")


            print(''.join([str(x) for x in encrypted_msg]))
            print("\nTotal time taken to encrypt:")
            #print(toc - tic)
            encrypted_salted_text = HashingAndSalting.salt(encrypted_msg)
            Sender_HASH = hashlib.sha256(encrypted_salted_text.encode()).hexdigest()
            wb=openpyxl.load_workbook("storedhashandsalt.xlsx")
            sh1=wb['hashed']
            sh1.cell(row=threading.active_count(),column=2,value=Sender_HASH)
            sh1.cell(row=threading.active_count(), column=1, value='192.168.56.1')
            sh1.cell(row=threading.active_count(), column=3, value=5050)
            #sh1.cell(row=threading.activeCount(), column=4,value=timeit.default_timer())
            sh1.cell(row=threading.active_count(), column=4, value=toc - tic)

            #sh1.cell(row=threading.active_count(), column=5, value=nm)

            wb.save("storedhashandsalt.xlsx")
            print("hashed value stored in database\n")

           # print("\nDecrypting message with private key ")
            #print("\nYour message is:")
           # tic = timeit.default_timer()
           # print(habsd.decrypt(encrypted_msg))
         #   toc = timeit.default_timer()
           # print("\nTotal time taken to decrypt:")
          #  print(toc - tic)
            conn.send("Msg received".encode(FORMAT))

    conn.close()


def start():
    server.listen()
    print(f"[LISTENING] Server is listening on {SERVER}")
    while True:
        conn, addr = server.accept()
        thread = threading.Thread(target=handle_client, args=(conn, addr))
        thread.start()
        print(f"[ACTIVE CONNECTIONS] {threading.active_count() - 1}")


print("[STARTING] server is starting...")
start()